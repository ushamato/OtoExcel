import io
import logging
import os
from typing import List, Tuple
from config import logger, SUPER_ADMIN_ID
from datetime import datetime
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import SQLAlchemyError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

class DatabaseManager:
    def __init__(self):
        # PostgreSQL bağlantı URL'si
        self.database_url = os.getenv('DATABASE_URL')
        if not self.database_url:
            raise ValueError("DATABASE_URL environment variable is not set")
        
        try:
            # SQLAlchemy engine oluştur
            self.engine = create_engine(
                self.database_url,
                pool_pre_ping=True,
                pool_size=5,
                max_overflow=10,
                pool_timeout=30
            )
            self.Session = sessionmaker(bind=self.engine)
            logger.info("Veritabanı bağlantısı başarıyla kuruldu")
            
        except Exception as e:
            logger.error(f"Veritabanı bağlantı hatası: {str(e)}")
            logger.error("Traceback:", exc_info=True)
            raise
    
    def setup_database(self):
        """Veritabanını kur"""
        try:
            print("Veritabanı kurulumu başlıyor...")
            # Senkron bağlantı kullan
            with self.engine.connect() as conn:
                print("Veritabanı bağlantısı başarılı!")
                # pgcrypto uzantısını etkinleştir ve commit et
                print("pgcrypto uzantısını yüklemeye çalışıyor...")
                conn.execute(text("CREATE EXTENSION IF NOT EXISTS pgcrypto;"))
                conn.commit()
                
                # Uzantının yüklendiğini doğrula
                print("pgcrypto uzantısını kontrol ediyor...")
                result = conn.execute(text("SELECT COUNT(*) FROM pg_extension WHERE extname = 'pgcrypto'"))
                if result.scalar() == 0:
                    logger.error("pgcrypto uzantısı yüklenemedi!")
                    print("pgcrypto uzantısı yüklenemedi!")
                    return False
                else:
                    logger.info("pgcrypto uzantısı başarıyla yüklendi.")
                    print("pgcrypto uzantısı başarıyla yüklendi.")
                
                # Gruplar tablosu
                print("Gruplar tablosunu oluşturuyor...")
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS groups (
                        id SERIAL PRIMARY KEY,
                        group_id BIGINT UNIQUE,
                        group_name TEXT,
                        added_by BIGINT,
                        added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """))
                
                # Admin tablosu
                print("Admin tablosunu oluşturuyor...")
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS group_admins (
                        user_id BIGINT PRIMARY KEY,
                        added_by BIGINT,
                        admin_name TEXT,
                        added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """))
                
                # Formlar tablosu
                print("Formlar tablosunu oluşturuyor...")
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS forms (
                        form_name TEXT,
                        group_id BIGINT,
                        fields TEXT,
                        created_by BIGINT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        PRIMARY KEY (form_name, group_id),
                        FOREIGN KEY (group_id) REFERENCES groups(group_id) ON DELETE CASCADE
                    )
                """))
                
                # Form gönderileri tablosu
                print("Form gönderileri tablosunu oluşturuyor...")
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS form_submissions (
                        id SERIAL PRIMARY KEY,
                        form_name TEXT,
                        group_id BIGINT,
                        user_id BIGINT,
                        chat_id BIGINT,
                        data TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (form_name, group_id) REFERENCES forms(form_name, group_id) ON DELETE CASCADE
                    )
                """))
                
                # Admin-Grup ilişki tablosu
                print("Admin-Grup ilişki tablosunu oluşturuyor...")
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS admin_groups (
                        admin_id BIGINT,
                        group_id BIGINT,
                        added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        PRIMARY KEY (admin_id, group_id),
                        FOREIGN KEY (admin_id) REFERENCES group_admins(user_id) ON DELETE CASCADE,
                        FOREIGN KEY (group_id) REFERENCES groups(group_id) ON DELETE CASCADE
                    )
                """))
                
                # Grup Bakiyeleri tablosu
                print("Grup Bakiyeleri tablosunu oluşturuyor...")
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS group_credits (
                        group_id BIGINT PRIMARY KEY,
                        credits FLOAT DEFAULT 0,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (group_id) REFERENCES groups(group_id) ON DELETE CASCADE
                    )
                """))
                
                # Admin Bakiyeleri tablosu
                print("Admin Bakiyeleri tablosunu oluşturuyor...")
                conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS admin_credits (
                        admin_id BIGINT PRIMARY KEY,
                        credits FLOAT DEFAULT 0,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (admin_id) REFERENCES group_admins(user_id) ON DELETE CASCADE
                    )
                """))
                
                conn.commit()
                logger.info("Veritabanı kurulumu tamamlandı.")
                print("Veritabanı kurulumu tamamlandı.")
                return True
        except Exception as e:
            logger.error(f"Veritabanı kurulum hatası: {str(e)}")
            print(f"Veritabanı kurulum hatası: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def get_groups(self, user_id=None):
        """Grupları getir"""
        try:
            if user_id is None:  # Süper admin
                query = """
                    SELECT g.group_id, g.group_name, g.id
                    FROM groups g
                    ORDER BY g.id
                """
                with self.engine.connect() as conn:
                    result = conn.execute(text(query))
                    return result.fetchall()
            else:  # Normal admin
                query = """
                    SELECT g.group_id, g.group_name, g.id
                    FROM groups g
                    INNER JOIN admin_groups ag ON g.group_id = ag.group_id
                    WHERE ag.admin_id = :user_id
                    ORDER BY g.id
                """
                with self.engine.connect() as conn:
                    result = conn.execute(text(query), {"user_id": user_id})
                    return result.fetchall()
        except Exception as e:
            logger.error(f"Grupları getirme hatası: {str(e)}")
            return None

    async def get_group_by_id(self, group_id):
        """Group ID ile grup bilgilerini getir"""
        try:
            query = """
                SELECT group_id, group_name
                FROM groups
                WHERE group_id = :group_id
            """
            with self.engine.connect() as conn:
                result = conn.execute(text(query), {"group_id": group_id})
                group = result.fetchone()
                if group:
                    return {
                        'group_id': group[0],
                        'group_name': group[1]
                    }
                return None
        except Exception as e:
            logger.error(f"Grup getirme hatası: {str(e)}")
            return None

    async def add_admin(self, user_id: str, admin_name: str = None, added_by: str = None) -> bool:
        """Admin ekle"""
        try:
            with self.engine.connect() as conn:
                result = conn.execute(text("""
                    INSERT INTO group_admins (user_id, added_by, admin_name)
                    VALUES (:user_id, :added_by, :admin_name)
                    ON CONFLICT (user_id) DO UPDATE
                    SET admin_name = EXCLUDED.admin_name
                """), {
                    "user_id": user_id,
                    "added_by": added_by,
                    "admin_name": admin_name
                })
                
                conn.commit()
                return True
        except SQLAlchemyError as e:
            logger.error(f"Admin ekleme DB hatası: {str(e)}")
            return False

    async def remove_admin(self, user_id: int) -> bool:
        """Admin sil"""
        try:
            with self.engine.connect() as conn:
                # Önce admin_groups tablosundan ilişkileri sil
                conn.execute(
                    text("""
                    DELETE FROM admin_groups
                    WHERE admin_id = :user_id
                    """),
                    {"user_id": user_id}
                )
                
                # Sonra admini sil
                conn.execute(
                    text("""
                    DELETE FROM group_admins
                    WHERE user_id = :user_id
                    """),
                    {"user_id": user_id}
                )
                
                conn.commit()
                return True
        except Exception as e:
            logger.error(f"Admin silme DB hatası: {str(e)}")
            return False

    async def get_all_admins(self) -> list:
        try:
            with self.engine.connect() as conn:
                # Telegram API'den admin isimlerini alamayız, bu yüzden veritabanında saklamamız gerekiyor
                # Şimdilik sadece ID'leri döndürüyoruz
                result = conn.execute(text("""
                    SELECT ga.user_id, ac.credits, ga.admin_name
                    FROM group_admins ga
                    LEFT JOIN admin_credits ac ON ac.admin_id = ga.user_id
                    ORDER BY ga.user_id
                """))
                
                admins = result.fetchall()
                return [
                    {
                        'user_id': admin[0],
                        'remaining_credits': admin[1] if admin[1] is not None else 0,
                        'admin_name': admin[2] if admin[2] is not None else "İsimsiz Admin"
                    }
                    for admin in admins
                ]
        except SQLAlchemyError as e:
            logger.error(f"Admin listeleme DB hatası: {str(e)}")
            return []

    async def Bakiye_ekle(self, admin_id: str, miktar: float) -> bool:
        """Admine Bakiye ekle"""
        try:
            with self.engine.connect() as conn:
                # Önce admin_id'nin var olup olmadığını kontrol et
                admin_check = conn.execute(text("""
                    SELECT COUNT(*) FROM group_admins 
                    WHERE user_id = :admin_id
                """), {"admin_id": admin_id}).scalar()
                
                if admin_check == 0:
                    logger.error(f"Admin bulunamadı: {admin_id}")
                    return False
                
                # Mevcut bakiyeyi al
                cursor = conn.execute(text("""
                    SELECT credits 
                    FROM admin_credits 
                    WHERE admin_id = :admin_id
                """), {"admin_id": admin_id})
                
                result = cursor.fetchone()
                current_balance = result[0] if result else 0
                
                # Yeni bakiyeyi hesapla ve güncelle
                new_balance = current_balance + miktar
                conn.execute(text("""
                    INSERT INTO admin_credits (admin_id, credits, updated_at)
                    VALUES (:admin_id, :new_balance, CURRENT_TIMESTAMP)
                    ON CONFLICT (admin_id) DO UPDATE 
                    SET credits = :new_balance, updated_at = CURRENT_TIMESTAMP
                """), {"admin_id": admin_id, "new_balance": new_balance})
                
                conn.commit()
                return True
        except SQLAlchemyError as e:
            logger.error(f"Bakiye ekleme DB hatası: {str(e)}")
            return False

    async def Bakiye_sil(self, admin_id: str, miktar: float) -> bool:
        """Adminden Bakiye sil"""
        try:
            with self.engine.connect() as conn:
                # Mevcut bakiyeyi al
                cursor = conn.execute(text("""
                    SELECT credits 
                    FROM admin_credits 
                    WHERE admin_id = :admin_id
                """), {"admin_id": admin_id})
                
                result = cursor.fetchone()
                if not result:
                    return False
                
                current_balance = result[0]
                if current_balance < miktar:
                    return False
                
                # Yeni bakiyeyi hesapla ve güncelle
                new_balance = current_balance - miktar
                conn.execute(text("""
                    UPDATE admin_credits 
                    SET credits = :new_balance, updated_at = CURRENT_TIMESTAMP
                    WHERE admin_id = :admin_id
                """), {"admin_id": admin_id, "new_balance": new_balance})
                
                conn.commit()
                return True
        except SQLAlchemyError as e:
            logger.error(f"Bakiye silme DB hatası: {str(e)}")
            return False

    async def bakiye_getir(self, admin_id: str) -> float:
        """Admin bakiyesini getir"""
        try:
            with self.engine.connect() as conn:
                cursor = conn.execute(text("""
                    SELECT credits 
                    FROM admin_credits 
                    WHERE admin_id = :admin_id
                """), {"admin_id": admin_id})
                
                result = cursor.fetchone()
                return result[0] if result else 0
        except SQLAlchemyError as e:
            logger.error(f"Bakiye getirme DB hatası: {str(e)}")
            return 0

    async def get_forms(self, user_id: int = None) -> list:
        try:
            with self.engine.connect() as conn:
                # Eğer user_id verilmişse, sadece o adminin formlarını getir
                if user_id:
                    cursor = conn.execute(text("""
                        SELECT form_name, fields 
                        FROM forms 
                        WHERE created_by = :user_id
                        ORDER BY form_name
                    """), {"user_id": user_id})
                # Super admin için tüm formları getir
                elif user_id == SUPER_ADMIN_ID:
                    cursor = conn.execute(text("""
                        SELECT form_name, fields 
                        FROM forms
                        ORDER BY form_name
                    """))
                # user_id verilmemişse boş liste dön
                else:
                    return []
                
                forms = cursor.fetchall()
                return [{'form_name': form[0], 'fields': form[1]} for form in forms]
        except SQLAlchemyError as e:
            logger.error(f"Form getirme DB hatası: {str(e)}")
            return []

    async def is_admin(self, user_id: int) -> bool:
        """Kullanıcının admin olup olmadığını kontrol et"""
        try:
            with self.engine.connect() as conn:
                cursor = conn.execute(text("""
                    SELECT COUNT(*) FROM group_admins 
                    WHERE user_id = :user_id
                """), {"user_id": user_id})
                count = cursor.scalar()
                return count > 0
        except SQLAlchemyError as e:
            logger.error(f"Admin kontrolü DB hatası: {str(e)}")
            return False

    async def is_group_admin(self, user_id: int) -> bool:
        try:
            with self.engine.connect() as conn:
                cursor = conn.execute(text("""
                    SELECT COUNT(*) FROM group_admins 
                    WHERE user_id = :user_id
                """), {"user_id": user_id})
                count = cursor.fetchone()[0]
                return count > 0
        except SQLAlchemyError as e:
            logger.error(f"Grup admin kontrolü DB hatası: {str(e)}")
            return False

    async def get_admin_groups(self, user_id: int) -> list:
        try:
            with self.engine.connect() as conn:
                cursor = conn.execute(text("""
                    SELECT g.group_id, g.group_name 
                    FROM group_admins ga 
                    JOIN groups g ON ga.group_id = g.group_id 
                    WHERE ga.user_id = :user_id
                """), {"user_id": user_id})
                groups = cursor.fetchall()
                return [{'id': group[0], 'name': group[1]} for group in groups]
        except SQLAlchemyError as e:
            logger.error(f"Admin grupları getirme DB hatası: {str(e)}")
            return []

    async def get_form(self, form_name, admin_id=None):
        """Form bilgilerini getir"""
        try:
            # SQL sorgusu
            if admin_id:
                query = text("""
                    SELECT form_name, fields, created_by
                    FROM forms
                    WHERE form_name = :form_name AND created_by = :admin_id
                """)
                params = {"form_name": form_name, "admin_id": admin_id}
            else:
                query = text("""
                    SELECT form_name, fields, created_by
                    FROM forms
                    WHERE form_name = :form_name
                """)
                params = {"form_name": form_name}
            
            # Sorguyu çalıştır
            with self.engine.connect() as conn:
                result = conn.execute(query, params)
                row = result.fetchone()
                
            if row:
                fields = row[1].split(',')
                return {
                    'form_name': row[0],
                    'fields': fields,
                    'created_by': row[2]
                }
            return None
            
        except Exception as e:
            logger.error(f"Form bilgisi getirme hatası: {str(e)}")
            return None

    async def check_duplicate_submission(self, form_name: str, group_id: int, data: str) -> bool:
        """Form verisinin daha önce kaydedilip kaydedilmediğini kontrol et"""
        try:
            # .env dosyasından şifreleme anahtarını al
            encryption_key = os.environ.get("POSTGRES_ENCRYPTION_KEY", "default_key_for_development")
            
            with self.engine.connect() as conn:
                # Önce form ve grup ID'sinin var olduğunu kontrol et
                check_query = text("""
                    SELECT COUNT(*) FROM forms 
                    WHERE form_name = :form_name AND group_id = :group_id
                """)
                check_result = conn.execute(check_query, {
                    "form_name": form_name,
                    "group_id": group_id
                })
                
                if check_result.scalar() == 0:
                    logger.error(f"Form bulunamadı: {form_name}, group_id: {group_id}")
                    return False
                
                # Şifre çözme işleminde güvenli parametreli sorgu kullan
                # cast fonksiyonu ile tip dönüşümlerini güvenli şekilde yap
                query = text("""
                    SELECT COUNT(*) 
                    FROM form_submissions 
                    WHERE form_name = :form_name 
                    AND group_id = :group_id 
                    AND cast(pgp_sym_decrypt(cast(data as bytea), cast(:encryption_key as text)) as text) = :data
                """)
                
                result = conn.execute(query, {
                    "form_name": form_name,
                    "group_id": group_id,
                    "data": data,
                    "encryption_key": encryption_key
                })
                
                count = result.scalar()
                return count > 0
        except SQLAlchemyError as e:
            logger.error(f"Mükerrer kayıt kontrolü DB hatası: {str(e)}")
            return False

    async def save_form_data(self, form_name: str, group_id: int, user_id: int, chat_id: int, data: str) -> int:
        """Form verisini kaydet ve submission_id'yi döndür"""
        try:
            # .env dosyasından şifreleme anahtarını al (.env ve GitHub secrets'a eklenmesi gerekli)
            encryption_key = os.environ.get("POSTGRES_ENCRYPTION_KEY")
            
            # Şifreleme anahtarı kontrolü
            if not encryption_key:
                logger.error("POSTGRES_ENCRYPTION_KEY bulunamadı!")
                return None
            
            with self.engine.connect() as conn:
                # Önce form ve grup ID'sinin var olduğunu kontrol et
                check_query = text("""
                    SELECT COUNT(*) FROM forms 
                    WHERE form_name = :form_name AND group_id = :group_id
                """)
                check_result = conn.execute(check_query, {
                    "form_name": form_name,
                    "group_id": group_id
                })
                
                if check_result.scalar() == 0:
                    logger.error(f"Form bulunamadı: {form_name}, group_id: {group_id}")
                    return None
                
                # Şifreleme işleminde güvenli parametreli sorgu kullan
                # cast fonksiyonu ile tip dönüşümlerini güvenli şekilde yap
                query = text("""
                    INSERT INTO form_submissions (form_name, group_id, user_id, chat_id, data)
                    VALUES (:form_name, :group_id, :user_id, :chat_id, 
                            pgp_sym_encrypt(cast(:data as text), cast(:encryption_key as text)))
                    RETURNING id
                """)
                
                result = conn.execute(query, {
                    "form_name": form_name,
                    "group_id": group_id,
                    "user_id": user_id,
                    "chat_id": chat_id,
                    "data": data,
                    "encryption_key": encryption_key
                })
                
                conn.commit()
                submission_id = result.scalar()
                return submission_id
        except SQLAlchemyError as e:
            logger.error(f"Form verisi kaydetme DB hatası: {str(e)}")
            return None

    async def generate_report(self, form_name: str, admin_id: int = None, 
                             start_date: datetime = None, end_date: datetime = None, is_super_admin: bool = False) -> io.BytesIO:
        """Form verilerinden Excel raporu oluştur"""
        try:
            # .env dosyasından şifreleme anahtarını al
            encryption_key = os.environ.get("POSTGRES_ENCRYPTION_KEY", "default_key_for_development")
            
            with self.engine.connect() as conn:
                # Form şablonunu al
                if is_super_admin:
                    # Süper admin tüm formları görebilir
                    cursor = conn.execute(text("""
                        SELECT fields FROM forms 
                        WHERE form_name = :form_name
                    """), {"form_name": form_name})
                else:
                    # Normal admin sadece kendi formlarını görebilir
                    cursor = conn.execute(text("""
                        SELECT fields FROM forms 
                        WHERE form_name = :form_name AND created_by = :admin_id
                    """), {"form_name": form_name, "admin_id": admin_id})
                
                form = cursor.fetchone()
                if not form:
                    logger.error(f"Form bulunamadı: {form_name}")
                    return None
                
                fields = form[0].split(',')
                
                # Verileri al - parametreli sorgu kullan
                base_query = """
                    SELECT cast(pgp_sym_decrypt(cast(fs.data as bytea), cast(:encryption_key as text)) as text), 
                           fs.created_at, fs.id
                    FROM form_submissions fs
                """
                
                params = {"encryption_key": encryption_key, "form_name": form_name}
                
                if is_super_admin:
                    # Süper admin tüm verileri görebilir
                    query = base_query + " WHERE fs.form_name = :form_name"
                else:
                    # Normal admin sadece kendi formlarının verilerini görebilir
                    query = base_query + """
                        JOIN forms f ON fs.form_name = f.form_name
                        WHERE fs.form_name = :form_name AND f.created_by = :admin_id
                    """
                    params["admin_id"] = admin_id
                
                # Tarih filtrelemesi
                if not start_date and not end_date:
                    today = datetime.now().strftime('%Y-%m-%d')
                    query += " AND DATE(fs.created_at) = :today"
                    params["today"] = today
                elif start_date and end_date:
                    query += " AND fs.created_at BETWEEN :start_date AND :end_date"
                    params["start_date"] = start_date.strftime('%Y-%m-%d 00:00:00')
                    params["end_date"] = end_date.strftime('%Y-%m-%d 23:59:59')
                
                query += " ORDER BY fs.id ASC"
                cursor = conn.execute(text(query), params)
                submissions = cursor.fetchall()
                
                if not submissions:
                    logger.error("Veri bulunamadı")
                    return None
                
                # Excel dosyası oluştur
                wb = Workbook()
                ws = wb.active
                ws.title = form_name
                
                # Stil tanımlamaları
                header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')  # Koyu lacivert başlık
                header_font = Font(bold=True, color='FFFFFF')  # Beyaz yazı
                center_align = Alignment(horizontal='center')
                
                # Satır renkleri (daha yoğun pastel tonlar)
                row_colors = [
                    'E3EEFF',  # Yoğun açık mavi
                    'FFE6E3',  # Yoğun açık somon
                    'E3FFEB',  # Yoğun açık mint
                    'FFF0E3',  # Yoğun açık şeftali
                ]
                
                # Başlıkları ekle
                headers = ['Form No'] + fields + ['Tarih']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                
                # Verileri ekle
                for row_idx, submission in enumerate(submissions, 2):
                    data = submission[0].split('\n')
                    created_at = submission[1]
                    form_id = submission[2]
                    
                    # Satır arkaplan rengi
                    row_color = row_colors[(row_idx-2) % len(row_colors)]
                    row_fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
                    
                    # Form numarası
                    cell = ws.cell(row=row_idx, column=1, value=form_id)
                    cell.fill = row_fill
                    cell.alignment = center_align
                    
                    # Form verileri
                    for col_idx, value in enumerate(data, 2):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.fill = row_fill
                    
                    # Tarih
                    cell = ws.cell(row=row_idx, column=len(headers), value=created_at)
                    cell.fill = row_fill
                    cell.alignment = center_align
                
                # Sütun genişliklerini ayarla
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    if column == 'A':  # Form No
                        ws.column_dimensions[column].width = 8
                    elif column == ws.cell(row=1, column=len(headers)).column_letter:  # Tarih
                        ws.column_dimensions[column].width = 19
                    else:
                        adjusted_width = min(max(max_length + 2, 10), 50)
                        ws.column_dimensions[column].width = adjusted_width
                
                # Excel dosyasını kaydet
                excel_file = io.BytesIO()
                wb.save(excel_file)
                excel_file.seek(0)
                
                return excel_file
                
        except Exception as e:
            logger.error(f"Rapor oluşturma hatası: {str(e)}")
            return None

    async def add_group(self, group_id: int, group_name: str, admin_id: int = None) -> bool:
        try:
            with self.engine.connect() as conn:
                # Grubu ekle
                conn.execute(text("""
                    INSERT INTO groups (group_id, group_name)
                    VALUES (:group_id, :group_name)
                    ON CONFLICT (group_id) DO UPDATE 
                    SET group_name = :group_name
                """), {"group_id": group_id, "group_name": group_name})

                # Eğer admin_id verilmişse ve süper admin değilse, admin-grup ilişkisini ekle
                if admin_id and admin_id != SUPER_ADMIN_ID:
                    conn.execute(text("""
                        INSERT INTO admin_groups (admin_id, group_id)
                        VALUES (:admin_id, :group_id)
                        ON CONFLICT DO NOTHING
                    """), {"admin_id": admin_id, "group_id": group_id})
                
                conn.commit()
                return True
        except SQLAlchemyError as e:
            logger.error(f"Grup ekleme DB hatası: {str(e)}")
            return False

    async def remove_group(self, group_id: int, admin_id: int = None) -> bool:
        try:
            with self.engine.connect() as conn:
                # Süper admin tüm grupları silebilir
                if admin_id == SUPER_ADMIN_ID:
                    # Önce admin-grup ilişkilerini sil
                    conn.execute(text("""
                        DELETE FROM admin_groups
                        WHERE group_id = :group_id
                    """), {"group_id": group_id})
                    
                    # Sonra grubu sil
                    result = conn.execute(text("""
                        DELETE FROM groups
                        WHERE group_id = :group_id
                    """), {"group_id": group_id})
                else:
                    # Normal admin sadece kendi grubunu silebilir
                    # Önce admin-grup ilişkisini kontrol et
                    cursor = conn.execute(text("""
                        SELECT COUNT(*) FROM admin_groups
                        WHERE admin_id = :admin_id AND group_id = :group_id
                    """), {"admin_id": admin_id, "group_id": group_id})
                    
                    if cursor.scalar() == 0:
                        return False
                    
                    # Admin-grup ilişkisini sil
                    conn.execute(text("""
                        DELETE FROM admin_groups
                        WHERE admin_id = :admin_id AND group_id = :group_id
                    """), {"admin_id": admin_id, "group_id": group_id})
                    
                    # Başka admin yoksa grubu da sil
                    cursor = conn.execute(text("""
                        SELECT COUNT(*) FROM admin_groups
                        WHERE group_id = :group_id
                    """), {"group_id": group_id})
                    
                    if cursor.scalar() == 0:
                        result = conn.execute(text("""
                            DELETE FROM groups
                            WHERE group_id = :group_id
                        """), {"group_id": group_id})
                    else:
                        # Başka adminler varsa grubu silme
                        return True
                
                conn.commit()
                return True
        except SQLAlchemyError as e:
            logger.error(f"Grup silme DB hatası: {str(e)}")
            return False

    async def get_group_name(self, group_id: int) -> str:
        try:
            with self.engine.connect() as conn:
                result = conn.execute(text("""
                    SELECT group_name 
                    FROM groups 
                    WHERE group_id = :group_id
                """), {"group_id": group_id})
                
                group = result.fetchone()
                return group[0] if group else None
        except SQLAlchemyError as e:
            logger.error(f"Grup adı getirme DB hatası: {str(e)}")
            return None

    async def get_form_submissions(self, form_name: str, group_id: int = None) -> list:
        try:
            with self.engine.connect() as conn:
                query = text("""
                    SELECT fs.id, fs.user_id, fs.chat_id, fs.data, fs.created_at
                    FROM form_submissions fs
                    WHERE fs.form_name = :form_name
                """)
                
                params = {"form_name": form_name}
                if group_id is not None:
                    query = text(query.text + " AND fs.group_id = :group_id")
                    params["group_id"] = group_id
                
                result = conn.execute(query, params)
                submissions = result.fetchall()
                
                return [
                    {
                        'id': sub[0],
                        'user_id': sub[1],
                        'chat_id': sub[2],
                        'data': sub[3],
                        'created_at': sub[4]
                    }
                    for sub in submissions
                ]
        except SQLAlchemyError as e:
            logger.error(f"Form gönderileri getirme DB hatası: {str(e)}")
            return []

    async def delete_form(self, form_name: str, group_id: int) -> bool:
        try:
            with self.engine.connect() as conn:
                # Önce form gönderilerini sil
                conn.execute(text("""
                    DELETE FROM form_submissions
                    WHERE form_name = :form_name AND group_id = :group_id
                """), {"form_name": form_name, "group_id": group_id})
                
                # Sonra formu sil
                result = conn.execute(text("""
                    DELETE FROM forms
                    WHERE form_name = :form_name AND group_id = :group_id
                """), {"form_name": form_name, "group_id": group_id})
                
                conn.commit()
                return result.rowcount > 0
        except SQLAlchemyError as e:
            logger.error(f"Form silme DB hatası: {str(e)}")
            return False

    async def delete_submission(self, submission_id: int) -> bool:
        try:
            with self.engine.connect() as conn:
                result = conn.execute(text("""
                    DELETE FROM form_submissions
                    WHERE id = :submission_id
                """), {"submission_id": submission_id})
                
                conn.commit()
                return result.rowcount > 0
        except SQLAlchemyError as e:
            logger.error(f"Form gönderisi silme DB hatası: {str(e)}")
            return False

    async def is_authorized_group(self, group_id: int) -> bool:
        """Grup yetkili bir admin tarafından eklenmiş mi kontrol et"""
        try:
            with self.engine.connect() as conn:
                # Önce groups tablosunda var mı kontrol et
                result = conn.execute(text("""
                    SELECT EXISTS (
                        SELECT 1 FROM groups 
                        WHERE group_id = :group_id
                    )
                """), {"group_id": group_id})
                
                return result.scalar()
        except SQLAlchemyError as e:
            logger.error(f"Grup yetki kontrolü DB hatası: {str(e)}")
            return False

    async def get_group_admins(self, group_id: int) -> list:
        """Grubun adminlerini getir"""
        try:
            with self.engine.connect() as conn:
                result = conn.execute(text("""
                    SELECT ga.user_id
                    FROM group_admins ga
                    INNER JOIN admin_groups ag ON ga.user_id = ag.admin_id
                    WHERE ag.group_id = :group_id
                """), {"group_id": group_id})
                
                admins = result.fetchall()
                return [
                    {
                        'user_id': admin[0]
                    }
                    for admin in admins
                ]
        except SQLAlchemyError as e:
            logger.error(f"Grup adminleri getirme DB hatası: {str(e)}")
            return []

    async def add_form(self, form_name, fields, user_id, chat_id=None):
        """Form ekle"""
        try:
            # Virgülle ayrılmış alanlar haline getir
            fields_str = ','.join(fields) if isinstance(fields, list) else fields
            
            # SQL sorgusu
            query = text("""
                INSERT INTO forms (form_name, fields, created_by, group_id)
                VALUES (:form_name, :fields, :user_id, :group_id)
                RETURNING form_name
            """)
            
            # Sorguyu çalıştır
            with self.engine.connect() as conn:
                # Eğer chat_id verilmişse onu kullan, yoksa user_id'yi kullan
                group_id = chat_id if chat_id is not None else user_id
                
                result = conn.execute(
                    query,
                    {
                        "form_name": form_name,
                        "fields": fields_str,
                        "user_id": user_id,
                        "group_id": group_id
                    }
                )
                returned_form_name = result.scalar()
                conn.commit()
                
            return returned_form_name is not None
            
        except Exception as e:
            logger.error(f"Form ekleme hatası: {str(e)}")
            return False

    def get_group_by_db_id(self, db_id):
        """DB ID ile grup bilgilerini getir"""
        try:
            query = """
                SELECT id, group_id, group_name
                FROM groups
                WHERE id = :db_id
            """
            with self.engine.connect() as conn:
                result = conn.execute(text(query), {"db_id": db_id})
                group = result.fetchone()
                if group:
                    return {
                        'id': group[0],
                        'group_id': group[1],
                        'group_name': group[2]
                    }
                return None
        except Exception as e:
            logger.error(f"DB ID ile grup getirme hatası: {str(e)}")
            return None 